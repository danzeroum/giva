"""MontadorSaida (RF-30) — CSV cru e `.xlsx` com formatação.

Determinismo por construção (RNF-04 — reprocessar produz saída idêntica):

- **ordem de colunas fixa**, declarada num único lugar (`COLUNAS_ENRIQUECIMENTO`);
- **proveniência serializada com chaves ordenadas** (`sort_keys=True`);
- **Decimal em formato canônico** (feito nas etapas: '18.0', nunca '18.00');
- terminador de linha fixo (`\n`).

O artefato de saída (colunas originais + enriquecimento) é o conteúdo estável.
Metadados voláteis (lote_id, timestamps) ficam FORA do artefato — em `lote`.

`montar_xlsx` colore cada linha pelo farol do seu pior status
(`giva.decisao.statuses.pior_status`) — a mesma paleta que o `StatusBadge` do
front usa, lida do mesmo `statuses.json` (contrato testado em
`tests/test_contrato_statuses.py`), para que a prévia na tela e o arquivo
baixado nunca divirjam.

O disclaimer da categoria (RN2) é anexado como comentário no cabeçalho da
coluna `categoria_macro` do `.xlsx` — a categoria é sugestão operacional, não
classificação fiscal."""

from __future__ import annotations

import csv
import io
import json
import re
from collections.abc import Sequence
from typing import Any

from giva.categoria.categorizador import DISCLAIMER_CATEGORIA
from giva.decisao.statuses import farol_de, pior_status
from giva.pipeline.modelo import LinhaLote

_FAROL_ARGB: dict[str, str] = {
    "verde": "FFE6F4EA",
    "amarelo": "FFFFF4E0",
    "vermelho": "FFFBE7E7",
}

# Caracteres que Excel/LibreOffice interpretariam como início de fórmula. Um
# '-' seguido de espaço é marcador de descrição (Classif), não fórmula — só
# sinaliza '-' antes de dígito. Original é entrada do usuário, nunca confiável.
_FORMULA = re.compile(r"^[=+@]|^-\d")


def _texto_seguro(valor: str) -> str:
    """Neutraliza injeção de fórmula (CSV/Excel): um apóstrofo força a
    célula a ser lida como texto, nunca avaliada como fórmula ao abrir a
    planilha."""
    return f"'{valor}" if _FORMULA.match(valor) else valor


# Ordem fixa das colunas de enriquecimento (RF-30 / saída F5 do PRD GIVA).
# `aliquota_icms_interna` é a modal nominal (sem FECP — decisão GIVA §3).
COLUNAS_ENRIQUECIMENTO: list[str] = [
    "descricao_oficial_ncm",
    "status_ncm",
    "aliquota_icms_interna",
    "status_aliquota",
    "fonte_aliquota",
    "observacao_aliquota",
    "categoria_macro",
    "confianca_categorizacao",
    "status_descricao",
    "status_linha",
    "proveniencia",
]


def _proveniencia_json(linha: LinhaLote) -> str:
    if not linha.proveniencia:
        return ""
    return json.dumps(linha.proveniencia, sort_keys=True, ensure_ascii=False)


def _celula(linha: LinhaLote, coluna: str) -> str:
    if coluna == "proveniencia":
        return _proveniencia_json(linha)
    return linha.enriquecimento.get(coluna, "")


def montar_csv(colunas_originais: list[str], linhas: list[LinhaLote]) -> str:
    """Gera o CSV: colunas originais preservadas na ordem + enriquecimento."""
    buffer = io.StringIO()
    escritor = csv.writer(buffer, lineterminator="\n")
    escritor.writerow(colunas_originais + COLUNAS_ENRIQUECIMENTO)
    for linha in linhas:
        originais = [_texto_seguro(linha.originais.get(c, "")) for c in colunas_originais]
        enriquecimento = [_celula(linha, c) for c in COLUNAS_ENRIQUECIMENTO]
        escritor.writerow(originais + enriquecimento)
    return buffer.getvalue()


def _status_da_linha(linha: LinhaLote) -> str:
    chaves = (
        linha.enriquecimento.get("status_ncm"),
        linha.enriquecimento.get("status_aliquota"),
        linha.enriquecimento.get("status_descricao"),
        linha.status_linha,
    )
    return pior_status(chave for chave in chaves if chave)


def montar_xlsx(colunas_originais: list[str], linhas: list[LinhaLote]) -> bytes:
    """Gera o `.xlsx` (RF-31): mesmas colunas do CSV, com cada linha colorida
    pelo farol do seu pior status; disclaimer da categoria no cabeçalho (RN2)."""
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import PatternFill

    preenchimentos = {
        farol: PatternFill("solid", fgColor=argb) for farol, argb in _FAROL_ARGB.items()
    }

    cabecalho = colunas_originais + COLUNAS_ENRIQUECIMENTO
    pasta = Workbook()
    planilha = pasta.active
    planilha.append(cabecalho)

    # Disclaimer (RN2) como comentário no cabeçalho de `categoria_macro`.
    coluna_categoria = cabecalho.index("categoria_macro") + 1
    planilha.cell(row=1, column=coluna_categoria).comment = Comment(
        DISCLAIMER_CATEGORIA, "GIVA"
    )

    for linha in linhas:
        originais = [_texto_seguro(linha.originais.get(c, "")) for c in colunas_originais]
        enriquecimento = [_celula(linha, c) for c in COLUNAS_ENRIQUECIMENTO]
        planilha.append(originais + enriquecimento)
        preenchimento = preenchimentos.get(farol_de(_status_da_linha(linha)))
        if preenchimento is not None:
            for celula in planilha[planilha.max_row]:
                celula.fill = preenchimento

    buffer = io.BytesIO()
    pasta.save(buffer)
    return buffer.getvalue()


def construir_resumo(linhas: Sequence[LinhaLote]) -> dict[str, Any]:
    """Resumo agregado de um lote processado, gravado em `lote.resumo`.

    `ok`/`invalidas` são por `status_linha`; `verde`/`amarelo`/`vermelho`,
    `motivos` e `categorias` são pelo pior status de cada linha. `disclaimer`
    acompanha o resumo para a interface exibi-lo (RN2)."""
    contagem_farol = {"verde": 0, "amarelo": 0, "vermelho": 0}
    contagem_motivo: dict[str, int] = {}
    contagem_categoria: dict[str, int] = {}
    for linha in linhas:
        pior = _status_da_linha(linha)
        farol = farol_de(pior)
        contagem_farol[farol] = contagem_farol.get(farol, 0) + 1
        if pior != "ok":
            contagem_motivo[pior] = contagem_motivo.get(pior, 0) + 1
        categoria = linha.enriquecimento.get("categoria_macro")
        if categoria:
            contagem_categoria[categoria] = contagem_categoria.get(categoria, 0) + 1

    return {
        "linhas": len(linhas),
        "ok": sum(1 for linha in linhas if linha.status_linha == "ok"),
        "invalidas": sum(
            1 for linha in linhas if linha.status_linha == "entrada_invalida"
        ),
        "verde": contagem_farol["verde"],
        "amarelo": contagem_farol["amarelo"],
        "vermelho": contagem_farol["vermelho"],
        "motivos": sorted(
            ({"motivo": motivo, "linhas": n} for motivo, n in contagem_motivo.items()),
            key=lambda item: -item["linhas"],
        ),
        "categorias": sorted(
            (
                {"categoria": categoria, "linhas": n}
                for categoria, n in contagem_categoria.items()
            ),
            key=lambda item: -item["linhas"],
        ),
        "disclaimer": DISCLAIMER_CATEGORIA,
    }
