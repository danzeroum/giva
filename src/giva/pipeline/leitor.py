"""LeitorPlanilha (RF-01/RF-02) — entrada `.csv` e `.xlsx` (F1 do PRD).

Mapeia cabeçalhos de forma tolerante (acentos/variações) às 4 colunas mínimas
(NCM, Período, Descrição, UF). Falta de qualquer uma → rejeita o lote inteiro
com a coluna ausente nomeada (HU-01: não processa parcialmente). Preserva TODAS
as colunas originais, na ordem, para a saída.

`ler_planilha` despacha por conteúdo: um `.xlsx` é um zip (assinatura `PK\\x03\\x04`);
qualquer outra coisa é tratada como CSV de texto.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass

from giva.pipeline.modelo import LinhaLote

# canônico → variações aceitas de cabeçalho (comparadas sem acento/caixa)
_ALIASES: dict[str, frozenset[str]] = {
    "ncm": frozenset({"ncm", "codigo", "codigo ncm", "cod ncm"}),
    "periodo": frozenset({"periodo", "data", "competencia", "mes", "mes/ano", "referencia"}),
    "descricao": frozenset({"descricao", "produto", "desc", "descricao do produto"}),
    "uf": frozenset({"uf", "estado", "unidade federativa"}),
}

_ACENTOS = str.maketrans("áàâãéêíóôõúüç", "aaaaeeiooouuc")

# Assinatura de arquivo ZIP — todo `.xlsx` (Office Open XML) começa com ela.
_ASSINATURA_XLSX = b"PK\x03\x04"


class ColunaAusenteError(Exception):
    """Uma coluna mínima não foi encontrada — o lote é rejeitado inteiro."""

    def __init__(self, ausentes: list[str]) -> None:
        self.ausentes = ausentes
        super().__init__(
            f"Colunas mínimas ausentes: {', '.join(ausentes)}. "
            f"O lote não é processado parcialmente (HU-01)."
        )


@dataclass(frozen=True)
class Lote:
    colunas_originais: list[str]
    linhas: list[LinhaLote]


def _canonizar(cabecalho: str) -> str:
    return " ".join(cabecalho.strip().lower().translate(_ACENTOS).split())


def _mapear_colunas(colunas: list[str]) -> dict[str, str]:
    """Resolve {campo canônico → nome real da coluna}. Erra se faltar alguma."""
    mapa: dict[str, str] = {}
    for real in colunas:
        canon = _canonizar(real)
        for campo, aliases in _ALIASES.items():
            if canon in aliases and campo not in mapa:
                mapa[campo] = real
    ausentes = [campo for campo in _ALIASES if campo not in mapa]
    if ausentes:
        raise ColunaAusenteError(ausentes)
    return mapa


def _montar_lote(colunas_originais: list[str], registros: list[dict[str, str]]) -> Lote:
    """Constrói o Lote a partir das colunas e das linhas já em dict (col→texto).
    Compartilhado por CSV e XLSX — o mapeamento e a preservação das originais
    são idênticos nas duas entradas."""
    mapa = _mapear_colunas(colunas_originais)
    linhas: list[LinhaLote] = []
    for numero, registro in enumerate(registros, start=1):
        originais = {c: (registro.get(c) or "") for c in colunas_originais}
        linhas.append(
            LinhaLote(
                numero=numero,
                originais=originais,
                bruto_ncm=originais[mapa["ncm"]],
                bruto_periodo=originais[mapa["periodo"]],
                bruto_uf=originais[mapa["uf"]],
                bruto_descricao=originais[mapa["descricao"]],
            )
        )
    return Lote(colunas_originais=colunas_originais, linhas=linhas)


def ler_csv(texto: str) -> Lote:
    leitor = csv.DictReader(io.StringIO(texto))
    colunas = leitor.fieldnames
    if not colunas:
        raise ColunaAusenteError(list(_ALIASES))
    return _montar_lote(list(colunas), list(leitor))


def ler_xlsx(conteudo: bytes) -> Lote:
    """Lê a primeira aba de um `.xlsx`. Toda célula vira texto (preserva o NCM
    como texto; a normalização cuida do zero à esquerda). Aba vazia / sem
    cabeçalho → erro de colunas ausentes."""
    from openpyxl import load_workbook

    pasta = load_workbook(io.BytesIO(conteudo), data_only=True, read_only=True)
    planilha = pasta.active
    linhas_iter = planilha.iter_rows(values_only=True)
    try:
        cabecalho_bruto = next(linhas_iter)
    except StopIteration:
        raise ColunaAusenteError(list(_ALIASES)) from None
    colunas = [str(c).strip() if c is not None else "" for c in cabecalho_bruto]
    if not any(colunas):
        raise ColunaAusenteError(list(_ALIASES))
    registros = [
        {colunas[i]: ("" if v is None else str(v)) for i, v in enumerate(linha) if i < len(colunas)}
        for linha in linhas_iter
    ]
    return _montar_lote(colunas, registros)


def ler_planilha(conteudo: bytes, nome_arquivo: str = "") -> Lote:
    """Despacha por conteúdo: `.xlsx` (assinatura ZIP) → `ler_xlsx`; senão CSV.
    O nome do arquivo é só um reforço — a decisão é pela assinatura."""
    if conteudo.startswith(_ASSINATURA_XLSX) or nome_arquivo.lower().endswith(".xlsx"):
        return ler_xlsx(conteudo)
    return ler_csv(conteudo.decode("utf-8"))
